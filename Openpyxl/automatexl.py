from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font 

#=========================INSERTING DATA INTO A SHEET OUTOMATICALLY==========================================
data = {
      "Joe":{
            "math":75,
            "science":89,
            "english":98,
            "gym":67
      },
      "Bill":{
            "math":100,
            "science":76,
            "english":90,
            "gym":55
      },

    "Tim":{
            "math":81,
            "science":93,
            "english":74,
            "gym":90
      },
      "Sally":{
            "math":79,
            "science":99,
            "english":78,
            "gym":69
      },
      "Jane":{
             "math":44,
            "science":85,
            "english":96,
            "gym":100
      }
      
}

wb = Workbook()
ws=wb.active
wb.title = "Grades"

headings = ["Name"] + list(data['Joe'].keys())
ws.append(headings)

for person in data:
    grades = list(data[person].values())
    ws.append([person] + grades)

for col in range(2, len(data['Joe'])+2):#Joe length is 4, but the range is started from 2, you must add the 2 to the last 
    char = get_column_letter(col)
    #this assigns formula as value to the target cells
    ws[char + "7"] = f"=SUM({char + '2'}:{char + '6'})/{len(data)}"

#give the cell A7 to a string value
ws["A7"] = "AVERAGE"

#change headers font style (first rows)
for col in range(1,6):
    ws[get_column_letter(col) + '1'].font = Font(bold=True,color='0099CCFF')

wb.save("/Users/jamalabdullahi/Python Tutorial/Python 101/File Handling/Openpyxl/NewGrades.xlsx")