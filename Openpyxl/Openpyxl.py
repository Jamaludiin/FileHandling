import openpyxl
#import the following libraries to open existing workbook
from openpyxl import workbook,load_workbook

#load a worksbook
wb = load_workbook('/Users/jamalabdullahi/Python Tutorial/Python 101/File Handling/Openpyxl/Grades.xlsx')

#kn=now the active worksheet from the loaded workbook 
ws = wb.active
#display the running or active worsheet
print("The active sheet is ", ws)

#access cell values from the active sheet and print its value on the screen
print(ws['A1'].value)#this means print what is has writen in the A1 (A-COLUMN, 1-ROW)

#ASSIGN NEW VALUE TO THE SPECIED CELL
ws['A2'].value = "Jamal"
#YOU CAN ALSO ASSIGN VALUES WITHOUT USING .VALUE AS FOLLOWS, .VALUE IS USED WHEN YOU WANT TO ACCESS, BUT IS ALSO LEGAL TO DO
# ws['A2'] = "Jamal"

#the new value will not be assigned yet, thus you must save the workbook as follows
wb.save('/Users/jamalabdullahi/Python Tutorial/Python 101/File Handling/Openpyxl/Grades.xlsx')


#=========CREATING, LISTING AND CHANGING SHEET=============

#display all the sheets in the workbook
print(wb.sheetnames)
#You can also print specific worksheet like this, this is also way to activate the sheet you want
print(wb['Sheet2'])

#assign new value to the active sheet. Note here we did not activated anysheet, thus this will depend which sheet is active at the workook
#activate sheet3 and the asigning value will go to sheet 3
ws=wb['Sheet3']
ws['A1'].value = "Name"
wb.save('/Users/jamalabdullahi/Python Tutorial/Python 101/File Handling/Openpyxl/Grades.xlsx')

#CREATE NEW SHEET 
wb.create_sheet("Test")
print(wb.sheetnames)