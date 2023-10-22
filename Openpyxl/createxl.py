from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active

ws.title = "Data"

ws.append(["One","Two","Three","Four","Five",10])
ws.append(["One","Two","Three","Four","Five",10])
ws.append(["One","Two","Three","Four","Five",10])
ws.append(["One","Two","Three","Four","Five",10])
ws.append(["One",921])

wb.save("/Users/jamalabdullahi/Python Tutorial/Python 101/File Handling/Openpyxl/jamal.xlsx")


#display the running or active worsheet
print("The active sheet is ", ws)

#access cell values from the active sheet and print its value on the screen
print(ws['B1'].value)#this means print what is has writen in the A1 (A-COLUMN, 1-ROW)

#ASSIGN NEW VALUE TO THE SPECIED CELL, IF THERE IS AN EXISTING VALUE IN THE TARGET CELL IT WILL OVERWRITE 
ws['A2'] = "Jamal"
wb.save("/Users/jamalabdullahi/Python Tutorial/Python 101/File Handling/Openpyxl/jamal.xlsx")
print(ws['A2'].value)

print()
#==========ACCESS THE CELL VALUES THROUGH LOOPING THE CELL VALUE AS RANGE
for row in range(1,11):
    for col in range(1,7):#six columns will be refered
        chr = get_column_letter(col)
        print(ws[chr + str(row)])#THIS PRINTS THE CELL NAMES EE.G. A1, B1, C1 ETC
        print(ws[chr + str(row)].value)#THIS PRINTS THE CELL VALUES


#===================ALTERNATIVE WAY OF DOING THE ABOVE====================
"""
GOT ERROR IN THIS ALTERNATIVES
for row in range(1,11):
    for col in range(0,7):#six columns will be refered
        char = chr(65 + col)
        print(ws[chr + str(row)].value)
"""


print()
#==========ASSIGN EACH CELL ITS NAME e.g. A1 assign A1 VALUE===================
wb.create_sheet("REASIGN") 
ws=wb['REASIGN']#ACTIVATE THE REASIGN SHEET IN ORDER TO STORE VALUES

for row in range(1,11):
    for col in range(1,7):#six columns will be refered
        chr = get_column_letter(col)
        ws[chr + str(row)]=chr + str(row)#THIS ASSIGNS NEW VALUES TO THE TARGET CELLS AND ASSIGN VALUES AS ITS NAMES EE.G. A1, B1, C1 ETC
        print(ws[chr + str(row)].value)#THIS PRINTS THE CELL VALUES OF THE ACTIVATED SHEET
#MUST SAVE AFTER THE OPERATION
wb.save("/Users/jamalabdullahi/Python Tutorial/Python 101/File Handling/Openpyxl/jamal.xlsx")


#==========MERGIN CELLS==========
#NOTE THIS WILL MERGE THE SPECIEF CELLS AND DATA INSIDE WILL BE LOST EXECPT THE DATA FOR THE FIRST CELL OF THE RANGE
#YOU MUST ALSO MAKE SURE THE KIND OF SHEET IS ACTIVATE OR YOU WANT TO ACTIVATE
ws.merge_cells("A1:F1")#THE SPECIFIED RANGE MUST COVER RANGE OF ROWS OR COLUMNS.
#ws.merge_cells("A1:F7")#THE SPECIFIED RANGE MUST COVER RANGE OF ROWS OR COLUMNS.
#MUST SAVE AFTER THE OPERATION
wb.save("/Users/jamalabdullahi/Python Tutorial/Python 101/File Handling/Openpyxl/jamal.xlsx")

#=========unmerge the cels===========
#NOTE THE PREVIOUS LOST DATA WILL NOT BE RETURNED AFTER UNMERGING
ws.unmerge_cells("A1:F1")
wb.save("/Users/jamalabdullahi/Python Tutorial/Python 101/File Handling/Openpyxl/jamal.xlsx")

#============INSERT ROWS AND DELETE==========
ws.insert_rows(7) #inserts one empty row at 7s rows, he shifts the data down if there is one
ws.insert_rows(7) #since the row 7 already made this adds new row to the seond or 8s position
#============DELETE ROWS============
ws.delete_rows(7)#deletes the last inserted row of 7,
ws.delete_rows(9)#after deleting it will shiftup the next data, while data in the target rows will be deletd as well


wb.save("/Users/jamalabdullahi/Python Tutorial/Python 101/File Handling/Openpyxl/jamal.xlsx")

#============INSERT COLUMNS AND DELETE==========
ws.insert_cols(2)#inserts new column after the first column, or new colunm will be added before column 3
#NO DATA WILL BE DELETD IN THIS PROCESS, BECOUSE THE PREVIOUS DATA IN PREVIOUS COLUMN WILL BE SHIFTED TO RIGHT

#DELETE A COLUMN
ws.delete_cols(2)#the new added column will be deletd
#ws.delete_cols(1)#if you delet the first column all its data will be lost, and above colunm aded will be shifted left as it will become the first colunm

wb.save("/Users/jamalabdullahi/Python Tutorial/Python 101/File Handling/Openpyxl/jamal.xlsx")


#============COPYING AND MOVING COLUMNS==========
ws.move_range("C1:F9",rows=1,cols=2)#shifts one row from above and two columns from left, while data to be sshifted is located in the specified range C1:F9
wb.save("/Users/jamalabdullahi/Python Tutorial/Python 101/File Handling/Openpyxl/jamal.xlsx")





#====================================================================================================================
var_dictionaries = {
                        "name": "Jamal",
                        "job": "Student",
                        "race": "Somali",
                        "age": 29,
                        "":2
                   }

print()
print(var_dictionaries)#prints all the keys and items (both keys and values are items) or values

print(var_dictionaries["name"])
print("")

print("Key names using keys() method are:")
for i in var_dictionaries.keys():
    print(i)

print()
print("Key names without using any method are:")
for i in var_dictionaries:
    print(i)

#USING ITEMS() METHOD
print("")
print("Key names using items() method are:")
for i in var_dictionaries.items():
    print(i)


#EXAMPLE 2: PRINT ALL VALUES OF A PYTHON DICTIONARY USING FOR LOOP
print("")
print("Value items using values() method are:")
for i in var_dictionaries.values():
    print(i)

print("")
print("Value items using values() method are:")
for i in var_dictionaries:
    print(i,":", var_dictionaries[i])

computer_courses = {
                        "course_one" : {
                            "name" : "Concept of programming",
                            "credit" : 4
                        },
                        "course_two" : {
                            "name" : "Object oriented paradigm",
                            "credit" : 3
                        },
                        "course_three" : {
                            "name" : "Calculus",
                            "credit" : 6
                        },
                        "course_four" : {
                            "mame" : "Software architecture",
                            "credit" : 3
                        }
}

print("\n------------------------------")
print(computer_courses)

print("\n===============================")
for i in computer_courses:
    print(i)

print()
for j in computer_courses.keys() :#same output to the above
     print(j)

print()
for j in computer_courses.keys() :#same output to the above
     print(j,computer_courses[j])

print("\n*********************************")
for i in computer_courses:
    print(computer_courses[i])

print()
for i in computer_courses.values():#same output to the above
    print(i)

print("\n+++++++++++++++++++++++++++++++++")
for j in computer_courses.items() :
     for i in j:
       print(i)

print()
for j in computer_courses.items() :#same output to the above
  print(j)

print()
for j in computer_courses.keys() :#same output to the above
     print(j,computer_courses[j])

print()
print("##############################")
for j in computer_courses.values() :#same output to the above
     print(j)

print()
print("values option four")
for i in computer_courses.values():
    for j in i:
     print(j, ":",i[j])

print()
print("_+_+_+_+_+_+_+_+_+_+_+_+_+_+_+_+_+_+_+_+_+_+_+_+_+_+_+_")
for i in computer_courses:
     print(computer_courses[i].values())

print()
for i in computer_courses:
     print(i, ":", computer_courses[i].values())

print()
for i in computer_courses:
     print([i], ":", computer_courses[i].values())




     