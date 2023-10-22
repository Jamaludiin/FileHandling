from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font 

#=========================INSERTING DATA INTO A SHEET OUTOMATICALLY==========================================
# WHAT TO DO 
# INSTEAD OF INSERTING DATA READ THE DATA ALREADY EXISTED IN clean CC1 FILE AND PERFORM CALCULATIONS

#wb = Workbook()
wb = load_workbook('/Users/jamalabdullahi/Python Tutorial/Python 101/File Handling/Openpyxl/GREEDY.xlsx')

ws=wb.active



# variable declarations
num_of_cols =60 #starts 2 column, thus we need 4, (6-2=4)
num_of_rows =46 # means 45 test cases plus 1 header row
num_of_test_case = num_of_rows-1

# those starting 7
val_starts_7 =num_of_cols+1
val_ends_11=num_of_cols+5

#CALCULATE IF EACH CELL UNDER A COLUMN HAS FAULT
val_insert_N = num_of_rows+1
val_range_start_N=2
val_range_end_N=num_of_rows
j=2
position = 0
"""for col in range(val_range_start_N, num_of_cols-1):#TCP1 length is 4, but the range is started from 2, you must add the 2 to the last 
    char = get_column_letter(col)
    #this assigns formula as value to the target cells
    while ws[char + str(j)].value == None:
      print("The active sheet is ", position+1, ws[char + str(j)], ws[char + str(j)].value)
      position+=1
      j+=1
    ws[char + str(val_insert_N)] = position+1"""

for i in range (val_range_start_N,num_of_cols):
    char = get_column_letter(i)
    j=2
    position = 0
    while ws[char + str(j)].value == None:
        #if j < num_of_rows:
        print("The active sheet is ", position+1, ws[char + str(j)], ws[char + str(j)].value )
        j+=1
        position+=1
    ws[char + str(val_insert_N)] = position+1
    #to track the cell values and position
    ws[char + str(val_insert_N+3)] = f"The active sheet is  {position+1} {ws[char + str(j)], ws[char + str(j)].value}"


# total positions
ws[get_column_letter(num_of_cols+1) + str(val_insert_N)] = f"=SUM({get_column_letter(val_range_start_N) + str(val_insert_N)}:{get_column_letter(num_of_cols) + str(val_insert_N)})"
# tapfd
ws[get_column_letter(num_of_cols+1+1) + str(val_insert_N)] = f"=1-(({get_column_letter(num_of_cols+1)+ str(val_insert_N)})/({int(num_of_rows-1)*int(num_of_cols-2)})) + 1/(2*{int(num_of_rows-1)})"



"""while ws['C' + str(j)].value == None:
            print("The active sheet is ", position+1,ws['C' + str(j)], ws['C' + str(j)].value )
            j+=1
            position+=1
ws['C' + str(val_insert_N)] = position+1"""

"""# FPV of simpson 
#CALCULATE FPV of 1-D*n
j=num_of_cols+num_of_cols+num_of_cols-1
for col in range(1,num_of_rows+1):
    char = get_column_letter(15)
    char1=get_column_letter(j)
    char2=get_column_letter(12)
    ws[char1 + str(col)] = f"=SUM({char2 + str(col)}:{char + str(col)})"

#= f"=SUM({char + str(val_range_start_N)}:{char + str(val_range_end_N)})"
ws[get_column_letter(num_of_cols+num_of_cols+num_of_cols-1)+'1']= "FPV 1-D*n" 

"""


wb.save("/Users/jamalabdullahi/Python Tutorial/Python 101/File Handling/Openpyxl/GREEDY.xlsx")
