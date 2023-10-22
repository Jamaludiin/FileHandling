from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font 
#from openpyxl.utils.dataframe import dataframe_to_rows

#=========================INSERTING DATA INTO A SHEET OUTOMATICALLY==========================================
# WHAT TO DO 
# INSTEAD OF INSERTING DATA READ THE DATA ALREADY EXISTED IN clean CC1 FILE AND PERFORM CALCULATIONS

#wb = Workbook()
wb = load_workbook('/Users/jamalabdullahi/Python Tutorial/Python 101/File Handling/Openpyxl/VENDING MECHINE FM.xlsx')

ws=wb.active



# variable declarations
num_of_cols =69 #starts 2 column, thus we need 4, (6-2=4)
num_of_rows =9 # means 45 test cases plus 1 header row
num_of_test_case = num_of_rows-1

# those starting 7
val_starts_7 =num_of_cols+1
val_ends_11=num_of_cols+5

#CALCULATE IF EACH CELL UNDER A COLUMN HAS FAULT
val_insert_N = num_of_rows+1
val_range_start_N=2
val_range_end_N=num_of_rows
j=2
position = 0



for i in range (val_range_start_N,num_of_cols):
    char = get_column_letter(i)
    j=2
    position = 0
    while ws[char + str(j)].value == None:
        #if j < num_of_rows:
        print("The active sheet is ", position+1, ws[char + str(j)], ws[char + str(j)].value )
        j+=1
        position+=1
    ws[char + str(val_insert_N)] = position+1
    #to track the cell values and position
    ws[char + str(val_insert_N+3)] = f"The active sheet is  {position+1} {ws[char + str(j)], ws[char + str(j)].value}"


# total positions
ws[get_column_letter(num_of_cols+1) + str(val_insert_N)] = f"=SUM({get_column_letter(val_range_start_N) + str(val_insert_N)}:{get_column_letter(num_of_cols) + str(val_insert_N)})"
# tapfd
ws[get_column_letter(num_of_cols+1+1) + str(val_insert_N)] = f"=1-(({get_column_letter(num_of_cols+1)+ str(val_insert_N)})/({int(num_of_rows-1)*int(num_of_cols-2)})) + 1/(2*{int(num_of_rows-1)})"


#sort
#ws[get_column_letter(1) + str(num_of_rows+5)] = f"=SORT({get_column_letter(1) + str(1)}:{get_column_letter(num_of_cols) + str(num_of_rows)},{num_of_cols},-1)"


wb.save("/Users/jamalabdullahi/Python Tutorial/Python 101/File Handling/Openpyxl/VENDING MECHINE FM.xlsx")
