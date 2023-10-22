from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font 

#=========================INSERTING DATA INTO A SHEET OUTOMATICALLY==========================================
data = {
      "TCP1":{
            "Instraction":75,
            "branch":89,
            "line":98,
            "method":67
      },
      "TCP2":{
            "Instraction":100,
            "branch":76,
            "line":90,
            "gym":55
      },

    "TCP3":{
            "Instraction":81,
            "branch":93,
            "line":74,
            "method":90
      },
      "TCP4":{
            "Instraction":79,
            "branch":99,
            "line":78,
            "method":69
      },
      "TCP5":{
             "Instraction":44,
            "branch":85,
            "line":96,
            "method":100
      },
        "TCP6":{
            "Instraction":84,
            "branch":57,
            "line":73,
            "method":61
      },
        "TCP7":{
            "Instraction":50,
            "branch":0,
            "line":85,
            "method":63
      },
        "TCP8":{
            "Instraction":79,
            "branch":67,
            "line":56,
            "method":80
      },
        "TCP9":{
            "Instraction":55,
            "branch":34,
            "line":6,
            "method":23
      },
        "TCP10":{
            "Instraction":23,
            "branch":76,
            "line":32,
            "method":46
      },
        "TCP11":{
            "Instraction":24,
            "branch":44,
            "line":67,
            "method":32
      },
        "TCP12":{
            "Instraction":46,
            "branch":42,
            "line":14,
            "method":66
      },
        "TCP13":{
            "Instraction":72,
            "branch":40,
            "line":56,
            "method":23
      }
}

wb = Workbook()
ws=wb.active
wb.title = "Simpson code coverage"

headings = ["Test Cases"] + list(data['TCP1'].keys())
ws.append(headings)

for person in data:
    grades = list(data[person].values())
    ws.append([person] + grades)

# variable declarations
num_of_cols =6 #starts 2 column, thus we need 4, (6-2=4)
num_of_rows =14 # means 13 test cases plus 1 header row

# those starting 7
val_starts_7 =num_of_cols+1
val_ends_11=num_of_cols+5

#CALCULATE N
val_insert_N = num_of_rows+1
val_range_start_N=2
val_range_end_N=num_of_rows
for col in range(val_range_start_N, num_of_cols):#TCP1 length is 4, but the range is started from 2, you must add the 2 to the last 
    char = get_column_letter(col)
    #this assigns formula as value to the target cells
    ws[char + str(val_insert_N)] = f"=SUM({char + str(val_range_start_N)}:{char + str(val_range_end_N)})"
   # ws[char + "15"] = f"=SUM({char + '2'}:{char + str(len(data)+1)})/{len(data)}"

#give the cell A15 to a string value
ws["A"+str(val_insert_N)] = "N"




#CALCULATE n(n-1)
val_range_start_n=2
count =1
for k in range(num_of_rows):
 count+=1
 j=num_of_cols+1
 for col in range(val_range_start_n, num_of_cols):
    char = get_column_letter(col)
    char1=get_column_letter(j)
    ws[char1 + str(count)] = f"={char + str(count)}*({char + str(count)}-1)"
    j+=1

ws[get_column_letter(num_of_cols+1)+'1']= "n(n-1)" 
#ws["G1"] = "n(n-1)" 
ws[get_column_letter(num_of_cols+1)+'1'].font = Font(bold=True,color='00FF6600')
#ws['G2'] = f"={'B2'}*({'B2'}-1)"

#CALCULATE Σ n(n-1)
val_insert_sum_n = num_of_rows+1
val_range_start_n=2

for col in range(val_starts_7, val_ends_11):
    char = get_column_letter(col)
    ws[char + str(val_insert_sum_n)] = f"=SUM({char + str(val_range_start_n)}:{char + str(num_of_rows)})"

ws[get_column_letter(val_starts_7-1)+str(val_insert_sum_n)] = "Σ n(n-1)" 

#CALCULATE N(N-1)
col2=2
for col in range(val_starts_7, val_ends_11):
    char = get_column_letter(col)
    char2 = get_column_letter(col2)
    ws[char + str(num_of_rows+2)] = f"={char2 + str(num_of_rows+1)}*({char2 + str(num_of_rows+1)}-1)"
    col2+=1

ws[get_column_letter(val_starts_7-1)+str(num_of_rows+2)] = "N(N-1)" 


# CALCULATE n(n-1)/N(N-1)
for col in range(val_starts_7, val_ends_11):
    char = get_column_letter(col)
    ws[char + str(num_of_rows+3)] = f"={char + str(num_of_rows+1)}/({char + str(num_of_rows+2)})"

ws[get_column_letter(val_starts_7-1)+str(num_of_rows+3)] = "n(n-1)/N(N-1)" 

# CALCULATE 1-D
for col in range(val_starts_7, val_ends_11):
    char = get_column_letter(col)
    ws[char + str(num_of_rows+4)] = f"=1-{char + str(num_of_rows+3)}"

ws[get_column_letter(val_starts_7-1)+str(num_of_rows+4)] = "1-D" 

# CALCULATE 1/D
for col in range(val_starts_7, val_ends_11):
    char = get_column_letter(col)
    ws[char + str(num_of_rows+5)] = f"=1/{char + str(num_of_rows+3)}"

ws[get_column_letter(val_starts_7-1)+str(num_of_rows+5)] = "1/D" 

#change headers font style (first rows)
F16_I16=val_starts_7-1

for col in range(1,val_starts_7):
    ws[get_column_letter(col) + '1'].font = Font(bold=True,color='00FF6600')
    ws[get_column_letter(col) + str(num_of_rows+1)].font = Font(bold=True,color='00FF6600')
    #ws[get_column_letter(col) + '16'].font = Font(bold=True,color='00FF6600')
    ws[get_column_letter(F16_I16-1)+str(num_of_rows+1)].font = Font(bold=True,color='00FF6600') 
    ws[get_column_letter(F16_I16-1)+str(num_of_rows+2)].font = Font(bold=True,color='00FF6600') 
    ws[get_column_letter(F16_I16-1)+str(num_of_rows+3)].font = Font(bold=True,color='00FF6600') 
    ws[get_column_letter(F16_I16-1)+str(num_of_rows+4)].font = Font(bold=True,color='00FF6600') 
    ws[get_column_letter(F16_I16)+str(num_of_rows+5)].font = Font(bold=True,color='00FF6600') 

    F16_I16+=1

print(len(data))
wb.save("/Users/jamalabdullahi/Python Tutorial/Python 101/File Handling/Openpyxl/Simpson_CC.xlsx")
